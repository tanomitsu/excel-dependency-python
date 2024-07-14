from __future__ import annotations
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import re
from typing import Set, List, TypeVar, Optional
import networkx as nx
import matplotlib.pyplot as plt
from pyvis.network import Network
from openpyxl.utils.cell import rows_from_range

T = TypeVar("T")

ADDRESS_PATTERN = re.compile(r"\b[A-Z]+[1-9][0-9]*(?::[A-Z]+[1-9][0-9]*)?\b")


# セルの数式から依存するセルアドレス一覧を取得する関数
def extract_cells_from_formula(formula) -> List[str]:
    addresses = []
    for match in ADDRESS_PATTERN.findall(formula):
        if ":" in match:
            # 範囲参照
            addresses.extend([cell for pair in rows_from_range(match) for cell in pair])
        else:
            addresses.append(match)

    return list(set(addresses))


# 1つのセルを表すクラス
class Cell:
    def __init__(self, sheet_name: str, address: str, value: Optional[float]):
        self.__sheet_name = sheet_name
        self.__address = address
        self.__value = value

    def sheet_name(self) -> str:
        return self.__sheet_name

    def address(self) -> str:
        return self.__address

    def value(self) -> Optional[float]:
        return self.__value


# セルの依存関係を表すクラス
class DependencyTree:
    def __init__(self, cell: "Cell"):
        self.__cell = cell
        self.__dependencies: Set[DependencyTree] = set()

    def add_dependency(self, parent: "DependencyTree") -> None:
        self.__dependencies.add(parent)

    def add_dependencies(self, parents: Set["DependencyTree"]) -> None:
        self.__dependencies.update(parents)

    def cell(self) -> "Cell":
        return self.__cell

    def dependencies(self) -> Set["DependencyTree"]:
        return self.__dependencies


# 配列からuniqueな値のみ取り出す関数
def unique_elements_preserve_order(input_list: List[T]) -> List[T]:
    return list(dict.fromkeys(input_list))


# セルアドレスをrootとして依存関係を計算
def get_value_or_function(workbook: "Workbook", root_cell_address: str) -> (str, str):
    sheet: Worksheet = workbook.active
    cell = sheet[root_cell_address]

    if cell.data_type == "f":
        root_cell = Cell(sheet_name="temp", address=root_cell_address, value=None)
        # Cell is function cell
        formula = cell.value
        parent_addresses = extract_cells_from_formula(formula)
        parent_addresses = unique_elements_preserve_order(parent_addresses)
        cur = DependencyTree(root_cell)
        for parent_address in parent_addresses:
            parent = get_value_or_function(workbook, parent_address)
            cur.add_dependency(parent)
        return cur
    else:
        # Cell is number cell
        root_cell = Cell(sheet_name="temp", address=root_cell_address, value=cell.value)
        return DependencyTree(root_cell)


def add_dependency_to_graph(node: "DependencyTree", G):
    edges = [(parent.cell().address(), node.cell().address()) for parent in node.dependencies()]
    G.add_edges_from(edges)
    for parent in node.dependencies():
        add_dependency_to_graph(parent, G)


def calc_dependency_graph(root: "DependencyTree") -> Network:
    # 有向グラフを作成
    g = nx.DiGraph()
    add_dependency_to_graph(root, g)
    net = Network(directed=True)
    net.from_nx(g)
    return net


def plot_dependency(root: "DependencyTree") -> None:
    net = calc_dependency_graph(root)
    net.show("output/dependency_tree.html", notebook=False)


def main():
    book_name = "testcases/" + input("Excel book name:")
    root_cell = input("Target cell address:")
    workbook: Workbook = load_workbook(book_name, data_only=False)
    root = get_value_or_function(workbook, root_cell)
    plot_dependency(root)


if __name__ == "__main__":
    main()
